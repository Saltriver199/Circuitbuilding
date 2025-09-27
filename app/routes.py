import os,subprocess,io
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, send_file, flash, session
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from .models import (db, Project, StationDrawing, JunctionBox, Circuit, 
                     Terminal, Group, TerminalHeader, ChokeTable, ResistorTable, get_ist_now)
from .schemas import SHEETS, HEADER_HINTS

bp = Blueprint("main", __name__)

# Model mapping for dynamic access based on sheet names
MODEL_MAP = {
    "StationDrawing": StationDrawing,
    "junction_box": JunctionBox,
    "circuit": Circuit,
    "terminal": Terminal,
    "group": Group,
    "terminal_header": TerminalHeader,
    "choketable": ChokeTable,
    "resistortable": ResistorTable,
}

def get_current_project():
    """Get current project from session WITHOUT auto-creating"""
    if 'project_id' not in session:
        return None  # Don't auto-create, return None
    
    project_id = session['project_id']
    # Verify the project still exists in database
    project = Project.query.get(project_id)
    if not project:
        # Project was deleted, clear from session
        session.pop('project_id', None)
        return None
    
    return project_id

def allowed_file(filename):
    """Check if uploaded file has allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

@bp.route("/")
def index():
    """Main page - redirect to project selection if no project"""
    project_id = get_current_project()
    
    if not project_id:
        # No project selected, redirect to project selection
        return redirect(url_for("main.project_selection"))
    
    current_project = Project.query.get(project_id)
    
    # Get row counts for each sheet
    table_counts = {}
    total_rows = 0
    for sheet_name, model in MODEL_MAP.items():
        count = model.query.filter_by(project_id=project_id).count()
        table_counts[sheet_name] = count
        total_rows += count
    
    return render_template("index.html", 
                         tables=table_counts, 
                         current_project=current_project,
                         total_rows=total_rows)

@bp.route("/project_selection")
def project_selection():
    """Project selection page - shows existing projects and create new option"""
    projects = Project.query.order_by(Project.created_date.desc()).all()
    
    # Get row counts for each project
    projects_data = []
    for project in projects:
        total_rows = 0
        for sheet_name, model in MODEL_MAP.items():
            count = model.query.filter_by(project_id=project.id).count()
            total_rows += count
        
        projects_data.append({
            'project': project,
            'total_rows': total_rows,
        })
    
    return render_template("project_selection.html", projects_data=projects_data)

@bp.route("/upload/<sheet_name>", methods=["GET", "POST"])
def upload_sheet(sheet_name):
    """Universal XLSX upload for any sheet"""
    if sheet_name not in SHEETS:
        flash(f"Unknown sheet: {sheet_name}")
        return redirect(url_for("main.index"))
    
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    
    current_project = Project.query.get(project_id)
    model = MODEL_MAP[sheet_name]
    expected_headers = SHEETS[sheet_name]
    
    if request.method == "POST":
        # Check if file was uploaded
        if 'file' not in request.files:
            flash('No file uploaded')
            return redirect(request.url)
        
        file = request.files['file']
        
        # Check if file was selected
        if file.filename == '':
            flash('No file selected')
            return redirect(request.url)
        
        # Check if file is allowed
        if not allowed_file(file.filename):
            flash('Only XLSX files are allowed')
            return redirect(request.url)
        
        try:
            # Load the workbook directly from memory
            wb = load_workbook(file, data_only=True)
            
            # Try to find sheet with matching name (case insensitive)
            sheet_found = None
            for ws_name in wb.sheetnames:
                if ws_name.lower() == sheet_name.lower():
                    sheet_found = ws_name
                    break
            
            if not sheet_found:
                flash(f'No "{sheet_name}" sheet found in the uploaded file. Available sheets: {", ".join(wb.sheetnames)}')
                return redirect(request.url)
            
            ws = wb[sheet_found]
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
                else:
                    headers.append('')
            
            # Check if required headers are present (case insensitive)
            missing_headers = []
            header_mapping = {}
            for required_header in expected_headers:
                found = False
                for i, file_header in enumerate(headers):
                    if file_header == required_header.lower():
                        header_mapping[required_header] = i
                        found = True
                        break
                if not found:
                    missing_headers.append(required_header)
            
            if missing_headers:
                flash(f'Missing required columns: {", ".join(missing_headers)}')
                return redirect(request.url)
            
            # Process rows and import data
            imported_count = 0
            error_count = 0
            
            # Start from row 2 (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Create data dictionary
                    data = {'project_id': project_id}
                    
                    # Map row values to expected headers
                    for header in expected_headers:
                        if header in header_mapping:
                            col_index = header_mapping[header]
                            # Get cell value, convert to string and handle None
                            cell_value = row[col_index] if col_index < len(row) else None
                            if cell_value is not None:
                                data[header] = str(cell_value).strip() if str(cell_value).strip() else None
                            else:
                                data[header] = None
                        else:
                            data[header] = None
                    
                    # Check if row has any data (not all None)
                    has_data = any(data[col] for col in expected_headers if data.get(col))
                    
                    if has_data:
                        # Create record for this sheet
                        record = model(**data)
                        db.session.add(record)
                        imported_count += 1
                        
                except Exception as e:
                    error_count += 1
                    print(f"Error processing row {row_num}: {str(e)}")
                    continue
            
            # Commit all changes
            if imported_count > 0:
                db.session.commit()
                flash(f'Successfully imported {imported_count} {sheet_name} records to Project ID {project_id}')
                if error_count > 0:
                    flash(f'Warning: {error_count} rows had errors and were skipped')
                return redirect(url_for("main.sheet_form", name=sheet_name))
            else:
                flash('No valid data found to import')
                return redirect(request.url)
                
        except Exception as e:
            db.session.rollback()
            flash(f'Error processing file: {str(e)}')
            return redirect(request.url)
    
    # GET request - show upload form
    return render_template("upload_sheet.html", 
                         current_project=current_project,
                         sheet_name=sheet_name,
                         expected_headers=expected_headers,
                         sheet_display_name=sheet_name.replace('_', ' ').title(),
                         hint=HEADER_HINTS.get(sheet_name, f"Upload {sheet_name} data from XLSX file"))

@bp.route("/sheet/<name>", methods=["GET", "POST"])
def sheet_form(name):
    """Handle form for individual sheets - add/edit/display"""
    if name not in SHEETS:
        flash(f"Unknown sheet: {name}")
        return redirect(url_for("main.index"))
    
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    
    current_project = Project.query.get(project_id)
    model = MODEL_MAP[name]
    columns = SHEETS[name]
    
    # Check if editing existing row
    edit_id = request.args.get('edit', type=int)
    edit_row = None
    if edit_id:
        edit_row = model.query.filter_by(id=edit_id, project_id=project_id).first()
        if not edit_row:
            flash("Row not found")
            return redirect(url_for("main.sheet_form", name=name))
    
    if request.method == "POST":
        # Collect form data
        data = {}
        for col in columns:
            value = request.form.get(col, "").strip()
            data[col] = value if value else None
        
        # Check if at least one field has data
        if any(data.values()):
            data['project_id'] = project_id
            
            if edit_id and edit_row:
                # Update existing row
                for col, val in data.items():
                    if hasattr(edit_row, col):
                        setattr(edit_row, col, val)
                flash(f"Row updated in {name} for Project ID {project_id}")
            else:
                # Create new row
                try:
                    new_record = model(**data)
                    db.session.add(new_record)
                    flash(f"Row added to {name} for Project ID {project_id}")
                except Exception as e:
                    flash(f"Error adding row: {str(e)}")
                    db.session.rollback()
                    return redirect(url_for("main.sheet_form", name=name))
            
            db.session.commit()
        else:
            flash("Please fill at least one field")
            
        return redirect(url_for("main.sheet_form", name=name))
    
    # GET request - display form and existing data
    rows = model.query.filter_by(project_id=project_id).order_by(model.id).all()
    
    # Convert SQLAlchemy objects to dictionaries for easier template access
    rows_data = []
    for row in rows:
        row_dict = {}
        for col in columns:
            row_dict[col] = getattr(row, col, '')
        row_dict['id'] = row.id  # Keep the ID for edit/delete
        rows_data.append(row_dict)
    
    # Convert edit_row to dictionary if exists
    edit_row_dict = None
    if edit_row:
        edit_row_dict = {}
        for col in columns:
            edit_row_dict[col] = getattr(edit_row, col, '')
    
    return render_template("sheet_form.html",
                         sheet=name, 
                         columns=columns, 
                         rows=rows_data,
                         hint=HEADER_HINTS.get(name, ""),
                         edit_id=edit_id,
                         edit_row=edit_row_dict,
                         current_project=current_project,
                         show_upload=True)  # Show upload button for ALL sheets

# ... (rest of the existing routes remain the same - delete_row, edit_row, preview, download, etc.)
@bp.route("/sheet/<name>/delete/<int:row_id>", methods=["POST"])
def delete_row(name, row_id):
    """Delete a specific row from database"""
    if name not in MODEL_MAP:
        flash("Invalid sheet name")
        return redirect(url_for("main.index"))
    
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    
    model = MODEL_MAP[name]
    
    record = model.query.filter_by(id=row_id, project_id=project_id).first()
    if record:
        try:
            db.session.delete(record)
            db.session.commit()
            flash(f"Row deleted successfully from {name} (Project ID: {project_id})")
        except Exception as e:
            flash(f"Error deleting row: {str(e)}")
            db.session.rollback()
    else:
        flash("Row not found")
    
    return redirect(url_for("main.sheet_form", name=name))

@bp.route("/sheet/<name>/edit/<int:row_id>")
def edit_row(name, row_id):
    """Redirect to sheet form with edit parameter"""
    return redirect(url_for("main.sheet_form", name=name, edit=row_id))

@bp.route("/preview")
def preview():
    """Preview all data before download"""
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    
    current_project = Project.query.get(project_id)
    
    table_data = {}
    total_records = 0
    for sheet_name, model in MODEL_MAP.items():
        records = model.query.filter_by(project_id=project_id).order_by(model.id).all()
        # Convert to dictionaries
        records_data = []
        for record in records:
            record_dict = {}
            for col in SHEETS[sheet_name]:
                record_dict[col] = getattr(record, col, '')
            records_data.append(record_dict)
        
        table_data[sheet_name] = records_data
        total_records += len(records_data)
    
    return render_template("preview.html", 
                         sheets=SHEETS, 
                         tables=table_data,
                         current_project=current_project,
                         total_records=total_records)

@bp.route("/download")
def download():
    """Generate and download XLSX file"""
    project_id = get_current_project()
    if not project_id:
        flash("Please select a project first")
        return redirect(url_for("main.project_selection"))
    
    current_project = Project.query.get(project_id)
    
    # Create workbook
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    total_records = 0
    # Add each sheet with data
    for sheet_name, columns in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        ws.append(columns)
        
        # Add data from database
        model = MODEL_MAP[sheet_name]
        records = model.query.filter_by(project_id=project_id).order_by(model.id).all()
        
        for record in records:
            row = []
            for col in columns:
                # Get attribute value, default to empty string
                value = getattr(record, col, "")
                row.append(str(value) if value is not None else "")
            ws.append(row)
            total_records += 1

    # Save to memory and send
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # Use IST time in filename
    filename = f"RAILWAYPROJECT_ID{project_id}_{current_project.name}_{get_ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    flash(f"Downloaded {total_records} records from Project ID {project_id}")
    
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@bp.route("/project/<int:project_id>/switch")
def switch_project(project_id):
    """Switch to different project"""
    project = Project.query.get_or_404(project_id)
    session['project_id'] = project.id
    flash(f"Switched to Project ID {project.id}: {project.name}")
    return redirect(url_for("main.index"))

@bp.route("/new_project", methods=["GET", "POST"])
def new_project():
    """Create new project"""
    if request.method == "POST":
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        
        if name:
            project = Project(name=name, description=description)
            db.session.add(project)
            db.session.commit()
            session['project_id'] = project.id
            flash(f"Created new Project ID {project.id}: {project.name}")
            return redirect(url_for("main.index"))
        else:
            flash("Project name is required")
    
    return render_template("new_project.html")

@bp.route("/clear_current_project", methods=["POST"])
def clear_current_project():
    """Clear all data from current project but keep project"""
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    
    try:
        total_deleted = 0
        # Delete all data for current project
        for model in MODEL_MAP.values():
            count = model.query.filter_by(project_id=project_id).count()
            model.query.filter_by(project_id=project_id).delete()
            total_deleted += count
        
        db.session.commit()
        flash(f"All data cleared from Project ID {project_id} (Deleted {total_deleted} records)")
    except Exception as e:
        flash(f"Error clearing project data: {str(e)}")
        db.session.rollback()
    
    return redirect(url_for("main.index"))

@bp.route("/excel_to_pdf", methods=["GET", "POST"])
def excel_to_pdf():
    """Upload XLSX and convert to PDF"""
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    
    current_project = Project.query.get(project_id)
    
    if request.method == "POST":
        # Check if file was uploaded
        if 'file' not in request.files:
            flash('No file uploaded')
            return redirect(request.url)
        
        file = request.files['file']
        
        # Check if file was selected
        if file.filename == '':
            flash('No file selected')
            return redirect(request.url)
        
        # Check if file is XLSX
        if not (file and allowed_file(file.filename)):
            flash('Only XLSX files are allowed')
            return redirect(request.url)
        
        try:
            # Create uploads directory if it doesn't exist
            upload_dir = os.path.join(os.getcwd(), 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Save uploaded file
            filename = secure_filename(file.filename)
            timestamp = get_ist_now().strftime('%Y%m%d_%H%M%S')
            xlsx_filename = f"railway_project_{project_id}_{timestamp}_{filename}"
            xlsx_path = os.path.join(upload_dir, xlsx_filename)
            file.save(xlsx_path)
            
            # Generate PDF filename
            pdf_filename = xlsx_filename.replace('.xlsx', '.pdf')
            pdf_path = os.path.join(upload_dir, pdf_filename)
            
            # Run the Excel to PDF converter script
            converter_script = os.path.join(os.getcwd(), 'excel_to_pdf_converter.py')
            
            # Execute the converter script
            result = subprocess.run([
                'python', converter_script, xlsx_path, pdf_path
            ], capture_output=True, text=True, timeout=300)  # 5 minute timeout
            
            if result.returncode == 0:
                flash(f'✅ Successfully converted {filename} to PDF!')
                # Clean up XLSX file
                os.remove(xlsx_path)
                
                return redirect(url_for('main.pdf_result', 
                                      filename=pdf_filename, 
                                      original_name=filename.replace('.xlsx', '.pdf')))
            else:
                flash(f'❌ Error converting file: {result.stderr}')
                # Clean up files on error
                if os.path.exists(xlsx_path):
                    os.remove(xlsx_path)
                return redirect(request.url)
                
        except subprocess.TimeoutExpired:
            flash('❌ Conversion timed out. File might be too large.')
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
            return redirect(request.url)
        except Exception as e:
            flash(f'❌ Error processing file: {str(e)}')
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
            return redirect(request.url)
    
    # GET request - show upload form
    return render_template("excel_to_pdf.html", current_project=current_project)

@bp.route("/pdf_result/<filename>/<original_name>")
def pdf_result(filename, original_name):
    """Show PDF result page with download option"""
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    
    current_project = Project.query.get(project_id)
    
    # Check if PDF file exists
    upload_dir = os.path.join(os.getcwd(), 'uploads')
    pdf_path = os.path.join(upload_dir, filename)
    
    if not os.path.exists(pdf_path):
        flash('PDF file not found')
        return redirect(url_for('main.excel_to_pdf'))
    
    return render_template("pdf_result.html", 
                         current_project=current_project,
                         filename=filename,
                         original_name=original_name)

@bp.route("/download_pdf/<filename>")
def download_pdf(filename):
    """Download the generated PDF"""
    upload_dir = os.path.join(os.getcwd(), 'uploads')
    pdf_path = os.path.join(upload_dir, filename)
    
    if not os.path.exists(pdf_path):
        flash('PDF file not found')
        return redirect(url_for('main.excel_to_pdf'))
    
    # Clean up file after download
    def remove_file(response):
        try:
            os.remove(pdf_path)
        except:
            pass
        return response
    
    return send_file(pdf_path, 
                    as_attachment=True, 
                    download_name=filename,
                    mimetype='application/pdf')